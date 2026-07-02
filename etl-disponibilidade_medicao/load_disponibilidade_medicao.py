"""
load_medicao_disponibilidade.py

Lê planilhas de Medição de Disponibilidade (layout largo: Equipes, FALTAS,
uma coluna por dia, Total Geral, VLR_HR), monta o modelo unpivot final e
consolida todos os arquivos de uma pasta de origem num único arquivo CSV.

Colunas do modelo final:
    arquivo_origem, tipo_equipe, Equipe, Faltas, Data, Qtd_Horas,
    Valor_Hora, Total_Geral
"""

import logging
import random
import re
import shutil
from pathlib import Path
from typing import List, Optional, Tuple

import openpyxl
import pandas as pd
from openpyxl.utils.exceptions import InvalidFileException

logger = logging.getLogger(__name__)

# ------------------------------------------------------------------
# ORDEM FINAL DAS COLUNAS
# ------------------------------------------------------------------
ORDEM_PADRAO = [
    'arquivo_origem',
    'tipo_equipe',
    'Equipe',
    'Faltas',
    'Data',
    'Qtd_Horas',
    'Valor_Hora',
    'Total_Geral',
]

# ------------------------------------------------------------------
# FUNÇÕES AUXILIARES DE TRATAMENTO
# ------------------------------------------------------------------

def _extrair_tipo_equipe(nome_arquivo: str) -> str:
    """Identifica a categoria macro do arquivo (ex: SEED MONEY, PERDAS, SMC)."""
    nome_maiusculo = nome_arquivo.upper()

    # Mantém a categorização macro conforme os nomes base
    if 'SEED' in nome_maiusculo:
        return 'SEED MONEY'
    elif 'PERDA' in nome_maiusculo:
        return 'PERDAS'
    elif 'SMC' in nome_maiusculo:
        return 'SMC'

    # Fallback para caso surja algum arquivo fora das 3 categorias previstas
    nome_limpo = re.sub(r'^\d+_\d+\s*', '', nome_arquivo)
    nome_limpo = re.sub(r'(?i)^Medição_Disponibilidade_', '', nome_limpo)
    return Path(nome_limpo).stem.strip()


def _localizar_linha_cabecalho(ws) -> int:
    """Retorna o índice (1-based) da linha que contém 'Equipes'."""
    for row in ws.iter_rows(min_row=1, max_row=10):
        for cell in row:
            if cell.value and str(cell.value).strip().lower() == 'equipes':
                return cell.row
    return -1


def _processar_aba(ws, nome_arquivo: str) -> Optional[pd.DataFrame]:
    """
    Processa uma aba individual e retorna o DataFrame no modelo unpivot.
    """
    tipo_equipe = _extrair_tipo_equipe(nome_arquivo)

    header_row = _localizar_linha_cabecalho(ws)
    if header_row == -1:
        logger.debug(f"Aba '{ws.title}' ignorada: cabeçalho 'Equipes' não localizado.")
        return None

    headers = [c.value for c in ws[header_row]]

    try:
        col_equipes = next(i for i, h in enumerate(headers) if str(h).strip().lower() == 'equipes')
        col_faltas = next(i for i, h in enumerate(headers) if str(h).strip().lower() == 'faltas')
    except StopIteration:
        logger.debug(f"Aba '{ws.title}' ignorada: colunas 'Equipes'/'Faltas' não localizadas.")
        return None

    col_datas = [i for i, h in enumerate(headers) if hasattr(h, 'strftime')]
    if not col_datas:
        logger.debug(f"Aba '{ws.title}' ignorada: nenhuma coluna de data encontrada.")
        return None

    # CORREÇÃO DA ANOMALIA DE DESLOCAMENTO (Meses < 31 dias)
    # A planilha padrão sempre reserva 31 colunas para os dias, independentemente
    # do mês. Portanto, as colunas finais estão a uma distância fixa do 1º DIA.
    primeira_col_data = col_datas[0]
    
    # +31 pula exatamente os 31 espaços de dias do template.
    col_total_geral = primeira_col_data + 31   # Refere-se à primeira coluna "Total Geral"
    col_vlr_hr = primeira_col_data + 32        # Refere-se à coluna "VLR_HR"

    registros = []
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        equipe = row[col_equipes].value
        if not equipe or not isinstance(equipe, str) or not equipe.startswith('PA-'):
            continue

        faltas = row[col_faltas].value
        total_geral = row[col_total_geral].value if col_total_geral < len(row) else None
        valor_hora = row[col_vlr_hr].value if col_vlr_hr < len(row) else None

        for col_idx in col_datas:
            data_valor = headers[col_idx]
            qtd_horas = row[col_idx].value

            registros.append({
                'arquivo_origem': nome_arquivo,
                'tipo_equipe': tipo_equipe,
                'Equipe': equipe,
                'Faltas': faltas,
                'Data': data_valor,
                'Qtd_Horas': qtd_horas,
                'Valor_Hora': valor_hora,
                'Total_Geral': total_geral,
            })

    if not registros:
        return None

    return pd.DataFrame(registros, columns=ORDEM_PADRAO)


def _processar_arquivo(caminho: Path, nome_saida: Optional[str] = None) -> Tuple[List[pd.DataFrame], bool]:
    """Abre o workbook e processa todas as abas visíveis."""
    dfs_extraidos = []
    arquivo_processado = False
    
    # Saneamento: Remover espaço irregular antes do .xlsx para consistência do DataFrame (arquivo_origem)
    nome_bruto = nome_saida or caminho.name
    nome_para_registro = re.sub(r'\s+(\.xlsx)$', r'\1', nome_bruto, flags=re.IGNORECASE)

    try:
        workbook = openpyxl.load_workbook(caminho, data_only=True)
    except (InvalidFileException, PermissionError, OSError) as e:
        logger.error(f"Erro ao abrir {caminho.name}: {e}")
        return [], False

    try:
        for ws in workbook.worksheets:
            if ws.sheet_state != 'visible':
                continue
            df_aba = _processar_aba(ws, nome_para_registro)
            if df_aba is not None and not df_aba.empty:
                dfs_extraidos.append(df_aba)
                arquivo_processado = True
    except Exception as e:
        logger.error(f"Erro inesperado no arquivo {caminho.name}: {e}", exc_info=True)
        return [], False
    finally:
        workbook.close()

    return dfs_extraidos, arquivo_processado


# ------------------------------------------------------------------
# API PÚBLICA
# ------------------------------------------------------------------

def montar_modelo(caminho_arquivo: Path, sheet_name: Optional[str] = None) -> pd.DataFrame:
    caminho_arquivo = Path(caminho_arquivo)
    wb = openpyxl.load_workbook(caminho_arquivo, data_only=True)
    
    nome_registro = re.sub(r'\s+(\.xlsx)$', r'\1', caminho_arquivo.name, flags=re.IGNORECASE)

    try:
        abas = [wb[sheet_name]] if sheet_name else [
            ws for ws in wb.worksheets if ws.sheet_state == 'visible'
        ]
        dfs = []
        for ws in abas:
            df_aba = _processar_aba(ws, nome_registro)
            if df_aba is not None and not df_aba.empty:
                dfs.append(df_aba)
    finally:
        wb.close()

    if not dfs:
        raise ValueError(f"Nenhum dado válido extraído de '{caminho_arquivo.name}'.")

    df = pd.concat(dfs, ignore_index=True)
    df.sort_values(by=['Equipe', 'Data'], inplace=True, ignore_index=True)
    return df


def processar_boletins(
    pasta_origem: Path,
    pasta_destino: Path,
    pasta_processados: Path,
) -> None:
    pasta_origem = Path(pasta_origem)
    pasta_destino = Path(pasta_destino)
    pasta_processados = Path(pasta_processados)

    pasta_destino.mkdir(parents=True, exist_ok=True)
    pasta_processados.mkdir(parents=True, exist_ok=True)

    logger.info("=" * 60)
    logger.info("CONSOLIDAÇÃO DE MEDIÇÕES DE DISPONIBILIDADE (CSV)")
    logger.info(f"Origem : {pasta_origem}")
    logger.info(f"Destino: {pasta_destino}")
    logger.info("=" * 60)

    arquivos = list(pasta_origem.rglob("*.xlsx"))
    arquivos = [f for f in arquivos if not f.name.startswith('~$')]

    # Ajustado para buscar consolidados no formato .csv
    consolidados = list(pasta_destino.glob("medicao-disponibilidade-consolidada_*.csv"))
    nomes_consolidados = {f.name for f in consolidados}
    arquivos = [f for f in arquivos if f.name not in nomes_consolidados]

    total_arquivos = len(arquivos)
    if total_arquivos == 0:
        logger.warning("Nenhum arquivo .xlsx novo encontrado para consolidação.")
        return

    logger.info(f"Encontrados {total_arquivos} arquivo(s) para processar.\n")

    todos_dfs = []
    arquivos_lidos = []
    arquivos_erro = []
    total_linhas = 0

    for idx, caminho in enumerate(arquivos, start=1):
        nome_saida = caminho.name # Mantém o nome original sempre

        logger.info(f"[{idx}/{total_arquivos}] Lendo: {caminho.name} ...")    
        dfs, sucesso = _processar_arquivo(caminho, nome_saida=nome_saida)
        if sucesso and dfs:
            total_linhas += sum(len(df) for df in dfs)
            todos_dfs.extend(dfs)
            arquivos_lidos.append((caminho, nome_saida))
            logger.info(f"  -> Sucesso: {len(dfs)} aba(s), {sum(len(df) for df in dfs)} linhas geradas após unpivot.")
        else:
            arquivos_erro.append(caminho)
            logger.warning("  -> Formato incompatível ou vazio. Mantido na origem.")

    if not todos_dfs:
        logger.error("Nenhum dado válido extraído de nenhum arquivo. Operação interrompida.")
        return

    logger.info("\nConcatenando todas as tabelas...")
    df_final = pd.concat(todos_dfs, ignore_index=True)
    df_final.sort_values(by=['Equipe', 'Data'], inplace=True, ignore_index=True)

    timestamp = pd.Timestamp.now().strftime('%Y%m%d%H%M%S')
    caminho_consolidado = pasta_destino / f'medicao-disponibilidade-consolidada_{timestamp}.csv'

    logger.info(f"Gravando arquivo consolidado: {caminho_consolidado.name} ...")
    try:
        # Passando o caminho completo do arquivo, e não apenas a pasta
        df_final.to_csv(caminho_consolidado, index=False, sep=';', encoding='utf-8-sig')
    except PermissionError:
        logger.error("ERRO: O arquivo final está bloqueado por outro processo ou sem permissão de escrita.")
        return

    total_salvas = len(df_final)

    logger.info("\n" + "=" * 60)
    logger.info("MÉTRICAS DO PROCESSAMENTO")
    logger.info("=" * 60)
    logger.info(f"Arquivos identificados      : {total_arquivos}")
    logger.info(f"Arquivos consolidados       : {len(arquivos_lidos)}")
    logger.info(f"Arquivos ignorados/erros    : {len(arquivos_erro)}")
    logger.info(f"Total de registros gerados  : {total_linhas}")
    logger.info(f"Total de registros salvos   : {total_salvas}")
    logger.info("-" * 60)

    if arquivos_erro:
        logger.warning("Arquivos que permaneceram na pasta de origem:")
        for arq in arquivos_erro:
            logger.warning(f"  - {arq.name}")

    if arquivos_lidos:
        logger.info(f"\nCopiando {len(arquivos_lidos)} arquivo(s) para a pasta de histórico...")
        copiados = 0
        for caminho_origem, nome_saida in arquivos_lidos:
            destino_arq = pasta_processados / nome_saida
            try:
                # Se o arquivo já existir na pasta de destino, exclui antes de copiar o novo
                if destino_arq.exists():
                    destino_arq.unlink()
                
                # Usa shutil.copy2 no lugar de shutil.move para preservar os metadados do arquivo
                shutil.copy2(str(caminho_origem), str(destino_arq))
                copiados += 1
            except Exception as e:
                logger.error(f"Falha ao copiar {caminho_origem.name}: {e}")
        logger.info(f"✅ {copiados} arquivo(s) copiados com sucesso.")

    logger.info("=" * 60)
    logger.info("EXECUÇÃO CONCLUÍDA COM SUCESSO!")