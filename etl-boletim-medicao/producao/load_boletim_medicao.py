import calendar
import logging
import re
import shutil
import unicodedata
from datetime import date
from pathlib import Path
from typing import List, Optional, Tuple

import random
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.table import Table, TableStyleInfo

logger = logging.getLogger(__name__)

# ------------------------------------------------------------------
# NORMALIZAÇÃO DO PERÍODO DE MEDIÇÃO (COMPETÊNCIA)
# Padrão de saída: dd.mm.yyyy a dd.mm.yyyy
#
# Promovido de homologacao/ para producao/ em 2026-09-02, depois de
# validado ponta a ponta contra o histórico bruto real de produção
# (ver CLAUDE.md/AGENTS.md, seção 8). Reconhece intervalo de data
# completo, "mm/yyyy" e mês por extenso (isolado ou em intervalo com
# dia, ex. "01 de Janeiro a 31 de Janeiro de 2026") -- mais abrangente
# que a versão anterior de producao/, que só reconhecia intervalo
# completo ou mês por extenso isolado (histórico: essa versão nasceu e
# evoluiu dentro de homologacao/, nunca em producao/).
# ------------------------------------------------------------------
_REGEX_PERIODO_MEDICAO = re.compile(
    r'(\d{1,2})[./-](\d{1,2})[./-](\d{2,4})\s*(?:a|à|-|até)\s*'
    r'(\d{1,2})[./-](\d{1,2})[./-](\d{2,4})',
    re.IGNORECASE,
)

# Formato "mm/yyyy" (mês e ano de referência, sem dia): expande para o
# período do mês inteiro (do dia 1 ao último dia do mês).
_REGEX_MES_ANO = re.compile(r'^(\d{1,2})[./-](\d{4})$')

# Mês por extenso (nome completo ou abreviado, sem acento).
_MESES_EXTENSO = {
    'janeiro': 1, 'jan': 1,
    'fevereiro': 2, 'fev': 2,
    'marco': 3, 'mar': 3,
    'abril': 4, 'abr': 4,
    'maio': 5, 'mai': 5,
    'junho': 6, 'jun': 6,
    'julho': 7, 'jul': 7,
    'agosto': 8, 'ago': 8,
    'setembro': 9, 'set': 9,
    'outubro': 10, 'out': 10,
    'novembro': 11, 'nov': 11,
    'dezembro': 12, 'dez': 12,
}
_ACENTOS = str.maketrans('áàâãéêíóôõúç', 'aaaaeeiooouc')
_REGEX_MES_NOME = re.compile(r'\b(' + '|'.join(_MESES_EXTENSO) + r')\b')
_REGEX_SEPARADOR_INTERVALO = re.compile(r'\s+(?:a|ate)\s+')
_REGEX_DIA_ISOLADO = re.compile(r'\b(\d{1,2})\b')
_REGEX_ANO_ISOLADO = re.compile(r'\b(\d{4})\b')


def _formatar_data_periodo(dia: str, mes: str, ano: str) -> str:
    ano_int = int(ano)
    if ano_int < 100:
        ano_int += 2000
    return f'{int(dia):02d}.{int(mes):02d}.{ano_int:04d}'


def _extrair_mes_extenso(texto_norm: str) -> Optional[int]:
    match = _REGEX_MES_NOME.search(texto_norm)
    return _MESES_EXTENSO[match.group(1)] if match else None


def _ano_referencia_para_mes(mes: int) -> int:
    """Mês por extenso sem ano: assume o ano atual, exceto se o mês for
    posterior ao mês corrente, caso em que se assume o ano anterior
    (a competência informada não pode estar no futuro)."""
    hoje = date.today()
    return hoje.year - 1 if mes > hoje.month else hoje.year


def _normalizar_periodo_mes_extenso(texto: str) -> Optional[str]:
    """Reconhece mês por extenso isolado ('Janeiro/2026') ou em intervalo
    início a fim ('01 de Janeiro a 31 de Janeiro de 2026')."""
    texto_norm = str(texto).lower().translate(_ACENTOS)
    if not _REGEX_MES_NOME.search(texto_norm):
        return None

    partes = _REGEX_SEPARADOR_INTERVALO.split(texto_norm, maxsplit=1)

    if len(partes) == 2:
        parte_inicio, parte_fim = partes

        mes_fim = _extrair_mes_extenso(parte_fim)
        ano_match_fim = _REGEX_ANO_ISOLADO.search(parte_fim)
        if mes_fim is None or not ano_match_fim:
            return None
        ano_fim = int(ano_match_fim.group(1))

        mes_inicio = _extrair_mes_extenso(parte_inicio) or mes_fim
        ano_match_inicio = _REGEX_ANO_ISOLADO.search(parte_inicio)
        ano_inicio = int(ano_match_inicio.group(1)) if ano_match_inicio else ano_fim

        dia_match_inicio = _REGEX_DIA_ISOLADO.search(parte_inicio)
        dia_inicio = int(dia_match_inicio.group(1)) if dia_match_inicio else 1

        dia_match_fim = _REGEX_DIA_ISOLADO.search(parte_fim)
        dia_fim = (
            int(dia_match_fim.group(1)) if dia_match_fim
            else calendar.monthrange(ano_fim, mes_fim)[1]
        )
    else:
        mes_inicio = mes_fim = _extrair_mes_extenso(texto_norm)
        if mes_inicio is None:
            return None
        ano_match = _REGEX_ANO_ISOLADO.search(texto_norm)
        if ano_match:
            ano_inicio = ano_fim = int(ano_match.group(1))
        else:
            ano_inicio = ano_fim = _ano_referencia_para_mes(mes_fim)
        dia_inicio = 1
        dia_fim = calendar.monthrange(ano_fim, mes_fim)[1]

    return (
        f'{_formatar_data_periodo(dia_inicio, mes_inicio, ano_inicio)} a '
        f'{_formatar_data_periodo(dia_fim, mes_fim, ano_fim)}'
    )


def _normalizar_periodo_medicao(valor):
    if pd.isna(valor):
        return valor
    texto = str(valor).strip()

    match = _REGEX_PERIODO_MEDICAO.search(texto)
    if match:
        d1, m1, a1, d2, m2, a2 = match.groups()
        padronizado = f'{_formatar_data_periodo(d1, m1, a1)} a {_formatar_data_periodo(d2, m2, a2)}'
        if padronizado != texto:
            logger.info(f"  Período de medição normalizado: '{texto}' -> '{padronizado}'")
        return padronizado

    match_mes_ano = _REGEX_MES_ANO.match(texto)
    if match_mes_ano:
        mes, ano = match_mes_ano.groups()
        mes_int = int(mes)
        ano_int = int(ano)
        ultimo_dia = calendar.monthrange(ano_int, mes_int)[1]
        padronizado = (
            f'{_formatar_data_periodo(1, mes_int, ano_int)} a '
            f'{_formatar_data_periodo(ultimo_dia, mes_int, ano_int)}'
        )
        logger.info(f"  Período de medição normalizado: '{texto}' -> '{padronizado}'")
        return padronizado

    padronizado_extenso = _normalizar_periodo_mes_extenso(texto)
    if padronizado_extenso:
        logger.info(f"  Período de medição normalizado: '{texto}' -> '{padronizado_extenso}'")
        return padronizado_extenso

    logger.info(f"  Período de medição não reconhecido, definido como nulo: '{texto}'")
    return pd.NA


# ------------------------------------------------------------------
# MAPEAMENTO COMPLETO (todas as variações → padrão)
# REPLICADO DE PRODUÇÃO (2026-09-01) -- chaves já normalizadas por
# _normalizar_cabecalho (sem acento/pontuação), não mais o
# .lower().replace(espaço) mais simples que a homologação usava.
# ------------------------------------------------------------------
MAPEAMENTO = {
    'distribuidora': 'distribuidora',
    'regional': 'regional',
    'tipodemedicao_cust_invest_ativ': 'tipo_medicao',
    'tipo_de_medicao_cust_invest_ativ': 'tipo_medicao',
    'tipomedicao': 'tipo_medicao',
    'estrutura_prod_disp': 'estrutura',
    'medicao_ciclo_final_pend': 'medicao',
    'periododemedicao': 'periodo_medicao',
    'periodo_de_medicao': 'periodo_medicao',
    'competencia': 'periodo_medicao',
    'competencia_servico': 'periodo_medicao',
    'parceiro': 'parceiro',
    'municipio': 'municipio',
    'municipiodolocaldaprestacaodeservico': 'municipio',
    'equipe': 'equipe',
    'descricaodanotafiscal': 'desc_nota_fiscal',
    'descricao_da_nota_fiscal': 'desc_nota_fiscal',
    'descricaonotafiscal': 'desc_nota_fiscal',
    'folhaderegistro': 'boletim',
    # 'folhregsrv' existia no MAPEAMENTO antigo da homologacao e tinha
    # sido perdido na reorganizacao de producao de 2026-09-01 -- gap real
    # (2 dos 6 arquivos que regrediram usam EXCLUSIVAMENTE esse
    # cabecalho abreviado pra boletim; o fallback heuristico
    # ('folha'+'registro') nao cobre a forma abreviada ('folh'+'regsrv')).
    # Corrigido nesta promocao (2026-09-02): a chave volta a existir em
    # producao/, deixando de ser um gap exclusivo de homologacao/.
    'folhregsrv': 'boletim',
    'boletim': 'boletim',
    'boletimdemedicao': 'boletim',
    'boletim_de_medicao': 'boletim',
    'valor': 'valor_bm',
    'datadeenviodoboletim': 'data_envio_bm',
    'data_de_envio_do_boletim': 'data_envio_bm',
    'cm': 'cp',
    'centro': 'cp',
    'cp': 'cp',
    'iva': 'iva',
    'nonotafiscal': 'nota_fiscal',
    'nf': 'nota_fiscal',
    'domiciliofiscal': 'domicilio_fiscal',
    'codigotarifafiscal': 'codigo_tarifa_fiscal',
    'cod_tarifafiscal': 'codigo_tarifa_fiscal',
    'identificadormedicao': 'identificador_medicao',
    'identificadoragrupamento': 'identificador_agrupamento',
    'contrato': 'contrato',
    'processo': 'processo',
    'textoboletim': 'texto_boletim',
    'coletorcusto': 'origem_lancamento',
    'origemdelancamento_pep_diagr_ord_cc': 'origem_lancamento',
    'origem_de_lancamento_pep_diagr_ord_cc': 'origem_lancamento',
    'notaproj': 'nota_prol',
}

# Colunas removidas do modelo final (não fazem mais parte do consolidado)
COLUNAS_REMOVIDAS = ['nota_fiscal', 'nota_prol']

# ------------------------------------------------------------------
# MAPEAMENTO HEURÍSTICO (fallback por combinação de palavras-chave,
# para cabeçalhos com variações de escrita não previstas no MAPEAMENTO
# exato acima — ex.: "Folha de Registro", "Registro/Folha" etc.)
# REPLICADO DE PRODUÇÃO (2026-09-01).
# ------------------------------------------------------------------
MAPEAMENTO_HEURISTICO = [
    (('folha', 'registro'), 'boletim'),
    (('pep', 'ordem'), 'origem_lancamento'),
]


def _mapear_por_heuristica(coluna: str) -> str:
    coluna = str(coluna)
    for palavras_chave, destino in MAPEAMENTO_HEURISTICO:
        if all(palavra in coluna for palavra in palavras_chave):
            return destino
    return coluna


# ------------------------------------------------------------------
# ORDEM FINAL DAS COLUNAS (a coluna "centro" não está nesta lista
# e será adicionada automaticamente no final)
# ------------------------------------------------------------------
ORDEM_PADRAO = [
    'distribuidora',
    'regional',
    'tipo_medicao',
    'estrutura',
    'medicao',
    'periodo_medicao',
    'parceiro',
    'municipio',
    'equipe',
    'desc_nota_fiscal',
    'boletim',
    'valor_bm',
    'data_envio_bm',
    'origem_lancamento',
    'cp',
    'iva',
    'domicilio_fiscal',
    'codigo_tarifa_fiscal',
    'identificador_medicao',
    'identificador_agrupamento',
    'contrato',
    'processo',
    'texto_boletim',
]


# ------------------------------------------------------------------
# FUNÇÕES AUXILIARES
# ------------------------------------------------------------------

def _normalizar_cabecalho(texto) -> str:
    texto = str(texto).strip().lower()
    texto = unicodedata.normalize('NFKD', texto).encode('ascii', 'ignore').decode('ascii')
    texto = re.sub(r'[^a-z0-9]+', '_', texto)
    return texto.strip('_')


def _localizar_linha_cabecalho(df_bruto: pd.DataFrame) -> int:
    for i, row in df_bruto.iterrows():
        linha_texto = ' '.join(row.fillna('').astype(str)).lower()
        if (
            'distribuidora' in linha_texto
            and 'regional' in linha_texto
            and 'valor' in linha_texto
        ):
            return i
    return -1


def _remover_linha_somatorio(df: pd.DataFrame) -> pd.DataFrame:
    if 'valor_bm' not in df.columns:
        return df
    outras_colunas = [c for c in df.columns if c != 'valor_bm']
    e_somatorio = df[outras_colunas].isna().all(axis=1) & df['valor_bm'].notna()
    return df[~e_somatorio]


def _processar_aba(sheet, nome_arquivo: str) -> Optional[pd.DataFrame]:
    dados_aba = list(sheet.iter_rows(values_only=True))
    if not dados_aba:
        return None

    df_bruto = pd.DataFrame(dados_aba)
    linha_cabecalho = _localizar_linha_cabecalho(df_bruto)
    if linha_cabecalho == -1:
        logger.debug(f"Aba '{sheet.title}' ignorada: cabeçalho não encontrado.")
        return None

    df_tabela = df_bruto.iloc[linha_cabecalho:].copy()
    df_tabela.columns = df_tabela.iloc[0].map(_normalizar_cabecalho)
    df_tabela = df_tabela.iloc[1:].copy()

    df_tabela.dropna(axis=1, how='all', inplace=True)
    df_tabela.dropna(axis=0, how='all', inplace=True)

    if df_tabela.empty:
        return None

    df_tabela = df_tabela.loc[:, ~df_tabela.columns.duplicated()]

    df_tabela.rename(columns=MAPEAMENTO, inplace=True)
    df_tabela.rename(columns=_mapear_por_heuristica, inplace=True)
    df_tabela = df_tabela.loc[:, ~df_tabela.columns.duplicated()]
    df_tabela.drop(columns=COLUNAS_REMOVIDAS, errors='ignore', inplace=True)

    if 'periodo_medicao' in df_tabela.columns:
        df_tabela['periodo_medicao'] = df_tabela['periodo_medicao'].apply(_normalizar_periodo_medicao)

    df_tabela = _remover_linha_somatorio(df_tabela)

    df_tabela['arquivo_origem'] = nome_arquivo

    colunas_validas = [
        col for col in df_tabela.columns
        if not str(col).lower().startswith('unnamed')
        and str(col).lower() not in ['nan', 'none', '', '<na>']
    ]
    df_tabela = df_tabela[colunas_validas]
    df_tabela = df_tabela.loc[:, ~df_tabela.columns.duplicated()]

    return df_tabela


def _processar_arquivo(caminho: Path, nome_saida: Optional[str] = None) -> Tuple[List[pd.DataFrame], bool]:
    dfs_extraidos = []
    arquivo_processado = False

    try:
        workbook = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    except (InvalidFileException, PermissionError, OSError) as e:
        logger.error(f"Erro ao abrir {caminho.name}: {e}")
        return [], False

    try:
        for sheet in workbook.worksheets:
            if sheet.sheet_state != 'visible':
                continue
            df_aba = _processar_aba(sheet, nome_saida or caminho.name)
            if df_aba is not None and not df_aba.empty:
                dfs_extraidos.append(df_aba)
                arquivo_processado = True
    except Exception as e:
        logger.error(f"Erro inesperado em {caminho.name}: {e}", exc_info=True)
        return [], False
    finally:
        workbook.close()

    return dfs_extraidos, arquivo_processado


# ------------------------------------------------------------------
# FUNÇÃO PRINCIPAL
# ------------------------------------------------------------------

def processar_boletins(
    pasta_origem: Path,
    pasta_destino: Path,
    pasta_processados: Optional[Path] = None,
    normalizar_deslocamentos: bool = False,
    caminho_auditoria: Optional[Path] = None,
    remover_sem_boletim: bool = False,
    excluir_boletim_nfse_duplicado: bool = False,
    remover_boletim_sem_valor_duplicado: bool = False,
) -> None:
    pasta_origem = Path(pasta_origem)
    pasta_destino = Path(pasta_destino)
    if pasta_processados is None:
        pasta_processados = pasta_origem / 'arquivos-processados'
    else:
        pasta_processados = Path(pasta_processados)

    pasta_destino.mkdir(parents=True, exist_ok=True)
    pasta_processados.mkdir(parents=True, exist_ok=True)

    logger.info("=" * 60)
    logger.info("CONSOLIDAÇÃO DE BOLETINS DE MEDIÇÃO")
    logger.info(f"Origem : {pasta_origem}")
    logger.info(f"Destino: {pasta_destino}")
    logger.info("=" * 60)

    arquivos = list(pasta_origem.glob("*.xlsx"))
    arquivos = [f for f in arquivos if not f.name.startswith('~$')]

    # Evita reprocessar consolidados existentes (CSV)
    consolidados = list(pasta_destino.glob("boletim-medicao-consolidado_*.xlsx"))
    nomes_consolidados = {f.name for f in consolidados}
    arquivos = [f for f in arquivos if f.name not in nomes_consolidados]

    total_arquivos = len(arquivos)
    if total_arquivos == 0:
        logger.warning("Nenhum arquivo .xlsx válido encontrado.")
        return

    logger.info(f"Encontrados {total_arquivos} arquivo(s).\n")

    todos_dfs = []
    arquivos_lidos = []
    arquivos_erro = []
    total_linhas = 0

    for idx, caminho in enumerate(arquivos, start=1):
        # Determina o nome de saída antes de processar:
        # se já existe arquivo igual em processados, o novo receberá prefixo aleatório
        destino_previsto = pasta_processados / caminho.name
        if destino_previsto.exists():
            prefixo = random.randint(1000, 9999)
            nome_saida = f"{prefixo} - {caminho.name}"
        else:
            nome_saida = caminho.name

        logger.info(f"[{idx}/{total_arquivos}] Processando: {caminho.name} ...")
        dfs, sucesso = _processar_arquivo(caminho, nome_saida=nome_saida)
        if sucesso and dfs:
            total_linhas += sum(len(df) for df in dfs)
            todos_dfs.extend(dfs)
            arquivos_lidos.append((caminho, nome_saida))
            logger.info(f"  -> OK: {len(dfs)} aba(s), {sum(len(df) for df in dfs)} linhas.")
        else:
            arquivos_erro.append(caminho)
            logger.warning(f"  -> Nenhum dado válido. Arquivo mantido na origem.")

    if not todos_dfs:
        logger.error("Nenhum dado extraído. Abortando.")
        return

    logger.info("\nConcatenando dados...")
    df_final = pd.concat(todos_dfs, ignore_index=True)

    # --- Ordenação final: todas as colunas padrão (mesmo ausentes no
    # lote), depois extras, arquivo_origem sempre por último ---
    # REPLICADO DE PRODUÇÃO (2026-09-01): antes, a homologação restringia
    # o consolidado só a ORDEM_PADRAO, descartando qualquer coluna extra
    # não mapeada -- achado real desta mesma sessão de homologação
    # (Coluna1/uf/id_ibge sendo silenciosamente perdidos). Produção já
    # preserva as extras via reindex; réplica exata adota o mesmo aqui.
    colunas_extras = [
        col for col in df_final.columns
        if col not in ORDEM_PADRAO and col != 'arquivo_origem'
    ]
    colunas_finais = ORDEM_PADRAO + colunas_extras
    if 'arquivo_origem' in df_final.columns:
        colunas_finais.append('arquivo_origem')
    df_final = df_final.reindex(columns=colunas_finais)

    if normalizar_deslocamentos:
        from normalizar_boletim import normalizar_dataframe

        logger.info("Aplicando normalização auditável de valores deslocados...")
        df_final, auditoria = normalizar_dataframe(df_final)
        if caminho_auditoria is None:
            timestamp_auditoria = pd.Timestamp.now().strftime('%Y%m%d%H%M%S')
            caminho_auditoria = pasta_destino / (
                f"boletim-medicao-auditoria_{timestamp_auditoria}.csv"
            )
        caminho_auditoria = Path(caminho_auditoria)
        caminho_auditoria.parent.mkdir(parents=True, exist_ok=True)
        auditoria.to_csv(caminho_auditoria, index=False, encoding='utf-8-sig')
        logger.info("Auditoria de normalização salva em: %s", caminho_auditoria)

    if remover_sem_boletim:
        # Grão do domínio é 1 linha = 1 boletim; registro sem número de
        # boletim não é um boletim de medição válido, independente de
        # quantas outras colunas ele tiver preenchidas (achado real na
        # homologação de 2026-08-31: 25.854 linhas nessa situação,
        # concentradas em arquivos de layout/domínio diferente do padrão).
        # Roda DEPOIS da normalização, pra dar chance de um boletim
        # genuinamente deslocado ser recuperado antes de descartar a linha.
        antes = len(df_final)
        sem_boletim = (
            df_final['boletim'].isna()
            | (df_final['boletim'].astype(str).str.strip() == '')
        )
        removidas = int(sem_boletim.sum())
        df_final = df_final[~sem_boletim].reset_index(drop=True)
        logger.info(
            "Removidas %d linha(s) sem boletim de %d (grão do domínio exige boletim).",
            removidas, antes,
        )

    if excluir_boletim_nfse_duplicado:
        # 'Boletim NFse.xlsx' tem qualidade de dado inferior às demais fontes
        # (achado real na homologação de 2026-08-31: 1.470 linhas-fantasma
        # com só o boletim preenchido, já assim na própria planilha de
        # origem). Quando o MESMO número de boletim já existe em qualquer
        # OUTRO arquivo_origem, a linha de 'Boletim NFse.xlsx' é descartada
        # em favor da outra fonte -- só permanece a linha dessa fonte quando
        # o boletim é exclusivo dela (nenhuma fonte melhor disponível).
        if 'arquivo_origem' in df_final.columns:
            antes = len(df_final)
            eh_nfse = df_final['arquivo_origem'].astype(str).str.endswith(
                'Boletim NFse.xlsx', na=False
            )
            boletins_outras_fontes = set(
                df_final.loc[~eh_nfse, 'boletim'].dropna()
            )
            duplicado_em_outra_fonte = eh_nfse & df_final['boletim'].isin(
                boletins_outras_fontes
            )
            removidas = int(duplicado_em_outra_fonte.sum())
            df_final = df_final[~duplicado_em_outra_fonte].reset_index(drop=True)
            logger.info(
                "Removidas %d linha(s) de 'Boletim NFse.xlsx' cujo boletim já "
                "existe em outra fonte.",
                removidas,
            )

    if remover_boletim_sem_valor_duplicado:
        # Generalização do achado acima: linha-fantasma (boletim
        # preenchido, valor_bm nulo) que é só duplicata incompleta de
        # OUTRA linha (mesmo boletim, com valor_bm preenchido) é
        # descartada -- independente do arquivo_origem (achado real na
        # homologação de 2026-09-01: 304 linhas em 200 boletins
        # distintos, espalhadas por 7 arquivos diferentes, inclusive
        # duplicata DENTRO do mesmo arquivo; 100% dos 200 boletins têm
        # valor recuperável em outra linha -- nenhum caso genuinamente
        # irrecuperável). Roda depois de excluir_boletim_nfse_duplicado
        # (que já remove Boletim NFse.xlsx por inteiro), então não há
        # sobreposição de critério.
        antes = len(df_final)
        boletins_com_valor = set(
            df_final.loc[df_final['valor_bm'].notna(), 'boletim'].dropna()
        )
        sem_valor_recuperavel = (
            df_final['boletim'].notna()
            & df_final['valor_bm'].isna()
            & df_final['boletim'].isin(boletins_com_valor)
        )
        removidas = int(sem_valor_recuperavel.sum())
        df_final = df_final[~sem_valor_recuperavel].reset_index(drop=True)
        logger.info(
            "Removidas %d linha(s)-fantasma (boletim sem valor) cujo "
            "boletim já existe em outra linha com valor preenchido.",
            removidas,
        )

    # --- Nome do arquivo com data e hora (yyyymmddhhmmss) ---
    timestamp = pd.Timestamp.now().strftime('%Y%m%d%H%M%S')
    caminho_consolidado = pasta_destino / f'boletim-medicao-consolidado_{timestamp}.xlsx'

    logger.info(f"Salvando consolidado em: {caminho_consolidado} ...")
    try:
        with pd.ExcelWriter(caminho_consolidado, engine='openpyxl') as writer:
            df_final.to_excel(writer, index=False, sheet_name='Boletins')
            planilha = writer.sheets['Boletins']
            ultima_coluna = get_column_letter(len(df_final.columns))
            ultima_linha = len(df_final) + 1
            tabela = Table(displayName='BoletinsMedicao', ref=f'A1:{ultima_coluna}{ultima_linha}')
            tabela.tableStyleInfo = TableStyleInfo(
                name='TableStyleMedium2', showRowStripes=True
            )
            planilha.add_table(tabela)
    except PermissionError:
        logger.error("ERRO: O arquivo destino está aberto em outro programa. Feche-o e tente novamente.")
        return

    total_salvas = len(df_final)

    # --- Relatório ---
    logger.info("\n" + "=" * 60)
    logger.info("RESUMO DA CONSOLIDAÇÃO")
    logger.info("=" * 60)
    logger.info(f"Arquivos na origem          : {total_arquivos}")
    logger.info(f"Arquivos processados        : {len(arquivos_lidos)}")
    logger.info(f"Arquivos com erro/sem dados : {len(arquivos_erro)}")
    logger.info(f"Linhas extraídas            : {total_linhas}")
    logger.info(f"Linhas salvas               : {total_salvas}")
    logger.info("-" * 60)

    if arquivos_erro:
        logger.warning("Arquivos NÃO processados (permanecem na origem):")
        for arq in arquivos_erro:
            logger.warning(f"  - {arq.name}")

    if arquivos_lidos:
        logger.info(f"\nMovendo {len(arquivos_lidos)} arquivo(s) para '{pasta_processados}' ...")
        movidos = 0
        for caminho_origem, nome_saida in arquivos_lidos:
            destino_arq = pasta_processados / nome_saida
            try:
                if nome_saida != caminho_origem.name:
                    logger.warning(
                        f"  ⚠️  Arquivo duplicado encontrado: '{caminho_origem.name}' "
                        f"já existia em processados. Movido como '{nome_saida}'."
                    )
                shutil.move(str(caminho_origem), str(destino_arq))
                movidos += 1
            except Exception as e:
                logger.error(f"Erro ao mover {caminho_origem.name}: {e}")
        logger.info(f"✅ {movidos} arquivo(s) movidos com sucesso.")
    else:
        logger.info("Nenhum arquivo para mover.")

    if total_linhas != total_salvas:
        logger.warning(
            f"⚠️ Diferença de linhas: extraídas={total_linhas}, salvas={total_salvas}. "
            "Possível eliminação de duplicatas ou linhas vazias."
        )
    else:
        logger.info("✅ Contagem de linhas consistente.")

    logger.info("=" * 60)
    logger.info("PROCESSAMENTO CONCLUÍDO!")


if __name__ == "__main__":
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s [%(levelname)s] %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    ORIGEM = Path(r'D:\One-Drive\Amper Elinsa\Fabiano Braz da Silva - Boletins de Medição\PARÁ')
    DESTINO = ORIGEM / 'base-consolidado'
    processar_boletins(ORIGEM, DESTINO)
